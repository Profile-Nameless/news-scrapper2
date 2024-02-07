from googlesearch import search
import openpyxl as pxl
import newspaper
import os
import pdfkit
import glob


def newssearch(query,language="en"):  # Function to search news articles related to the query

    try:   
        wb=pxl.load_workbook('newslinks.xlsx')
        try:   # If the worksheet exists, load it and get the existing URLs
            ws=wb[query+'newslink']
            existing_urls=[row[0].value for row in ws.iter_rows(min_row=2,min_col=1,max_col=1)]
        except: # If the worksheet does not exist, create it
            ws=wb.create_sheet(newsquery+'newslink')
            ws.append(['URL','Title','Description','Author','Publishing Date'])
            existing_urls=[]
    
    except:  #If the workbook does not exist, create it
        
        wb=pxl.Workbook()
        ws=wb.active
        ws.title=newsquery+'newslink'
        ws.append(['URL','Title','Description','Author','Publishing Date'])
        existing_urls=[]
    result=search(query,lang=language,num=10,extra_params={'tbm': 'nws'})
    data=[]
     # For each result, if it is not in the existing URLs, parse it and add it to the data
    for i in result:
        if len(data)>=10:
            break
        if i in existing_urls:
            continue
        try:
            print(i)
            article = newspaper.Article(i)
            article.download()
            article.parse()
            if article.publish_date:
                date_only = article.publish_date.date()  # Extracts only the date
                
            else:
                date_only="N A"
                    
            
            if article.authors:
                authors = ', '.join(article.authors)  # This joins all authors into a single string
            else:
                authors = "N A"
                
            if article.text:
                text=article.text
                
            else:
                text="N A"
                
            data.append([i,article.title, text,authors,date_only])
        except Exception as e:
            print("An error occured")
            continue
    return data
        
def createxl(newsquery):  # Function to create an Excel file with the news data

    data=newssearch(newsquery)
    try:   
        wb=pxl.load_workbook('newslinks.xlsx')
        try:
            ws=wb[newsquery+'newslink']
        except:
            ws=wb.create_sheet(newsquery+'newslink')
            ws.append(['URL','Title','Description','Author','Publishing Date'])
    
    except:
        
        wb=pxl.Workbook()
        ws=wb.active
        ws.title=newsquery+'newslink'
        ws.append(['URL','Title','Description','Author','Publishing Date'])
    
    finally:    
        
        for row in data:
                    ws.append(row)
        wb.save(os.path.join(os.getcwd(),'newslinks.xlsx'))
    return data
def createpdf(newsquery): # Function to create PDFs from the news data
    data = createxl(newsquery)
    output_directory = os.path.join(os.getcwd(),"Google_News_PDFs",newsquery)
    print(data)
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)
    
    existing_pdfs=glob.glob(os.path.join(output_directory,"*.pdf"))
    start_idx=len(existing_pdfs)  # Get the number of existing PDFs
    
    # Set path to wkhtmltopdf executable
    config = pdfkit.configuration(wkhtmltopdf=r"C:\Users\LENOVO\Downloads\wkhtmltox-0.12.6-1.mxe-cross-win64\bin\wkhtmltopdf.exe")

    # Convert each search result page to PDF
    for idx, link in enumerate(data,start=start_idx):
        print(link)
        try:
            print(f"Converting page {idx+1}: {link[0]}")
            pdfkit.from_url(link[0], os.path.join(output_directory, f"page_{idx+1}.pdf"), configuration=config)
            print(f"Page {idx+1} converted to PDF successfully!")
        except Exception as e:
            print(f"Failed to convert page {idx+1} to PDF: {str(e)}")


# Get the search query from the user
newsquery=input("Enter search query:")


createpdf(newsquery)
